from matplotlib import animation, use
from scipy.interpolate import interp1d
use("TkAgg")
import os
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import matplotlib.pyplot as plt
from matplotlib.animation import FuncAnimation
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from scipy.signal import lombscargle
import scipy.signal as sp_signal
import wfdb
import matplotlib.ticker as ticker
from scipy.signal import savgol_filter
import colorsys
import time
from scipy.integrate import trapezoid
from scipy.signal import welch, find_peaks
from scipy.stats import kurtosis, skew
from scipy import signal
import numpy as np
import pywt
from sklearn.linear_model import LogisticRegression


class SignalProcessor:
    def __init__(self, path):
        self._filter_cache = {}
        self.record_name = path
        self.record = wfdb.rdrecord(self.record_name)
        self.fs = self.record.fs
        self.original_ecg = self.record.p_signal[:, 0]
        self.current_signal = self.original_ecg.copy()
        self.window_size = int(3.0 * self.fs)
        self.step = 5
        self.index = 0
        self.is_running = True
        self.signal_level = 0.0
        self.noise_level = 0.0
        self.thresh_level = 0.0

        # pipeline:
        self.filters_pipeline = []
        self.filter_names = set()
        self._sos_states = {}

        #runtime state
        self.all_rr_times = []
        self.history_seconds = 10
        self.last_segment = None
        self.use_zero_phase = False

    def _get_sos(self, kind, fs, *params):
        try:
            key_params = []
            for p in params:
                if isinstance(p, (list, tuple, np.ndarray)):
                    key_params.extend([float(x) for x in p])
                else:
                    key_params.append(float(p))
            key = (kind, float(fs)) + tuple(key_params)
        except Exception as e:
            raise ValueError(f"Bad params for _get_sos: {params}") from e

        sos = self._filter_cache.get(key)
        if sos is not None:
            return sos, key

        nyq = 0.5 * float(fs)
        if kind == 'low':
            cutoff, order = params
            wn = float(cutoff) / nyq
            sos = sp_signal.butter(int(order), wn, btype='low', output='sos')
        elif kind == 'high':
            cutoff, order = params
            wn = float(cutoff) / nyq
            sos = sp_signal.butter(int(order), wn, btype='high', output='sos')
        elif kind == 'band':
            lowcut, highcut, order = params
            wn = [float(lowcut) / nyq, float(highcut) / nyq]
            sos = sp_signal.butter(int(order), wn, btype='band', output='sos')
        else:
            raise ValueError("unknown filter kind")

        self._filter_cache[key] = sos
        return sos, key

    def low(self, sig, fs, cutoff, order=5, zero_phase=False, streaming=True, key=None):
        sos, internal_key = self._get_sos('low', fs, float(cutoff), int(order))
        if sos is None:
            return sig
        used_key = key if key is not None else internal_key
        if streaming and not zero_phase:
            return self._sos_filter_stream(sos, sig, used_key)
        elif zero_phase:
            return sp_signal.sosfiltfilt(sos, sig)
        else:
            return sp_signal.sosfilt(sos, sig)

    def high(self, sig, fs, cutoff, order=5, zero_phase=False, streaming=True, key=None):
        sos, internal_key = self._get_sos('high', fs, float(cutoff), int(order))
        if sos is None:
            return sig
        used_key = key if key is not None else internal_key
        if streaming and not zero_phase:
            return self._sos_filter_stream(sos, sig, used_key)
        elif zero_phase:
            return sp_signal.sosfiltfilt(sos, sig)
        else:
            return sp_signal.sosfilt(sos, sig)

    def band(self, sig, fs, lowcut, highcut, order=5, zero_phase=False, streaming=True, key=None):
        sos, internal_key = self._get_sos('band', fs, float(lowcut), float(highcut), int(order))
        if sos is None:
            return sig
        used_key = key if key is not None else internal_key
        if streaming and not zero_phase:
            return self._sos_filter_stream(sos, sig, used_key)
        elif zero_phase:
            return sp_signal.sosfiltfilt(sos, sig)
        else:
            return sp_signal.sosfilt(sos, sig)

    def compute_spectrum(self, sig):
        import numpy as np
        from scipy.signal import welch
        x = np.asarray(sig, dtype=float)
        if x.size < 8:
            return np.array([0.0]), np.array([1e-12])
        x = x - np.mean(x)
        nperseg = min(1024, max(64, x.size // 2))
        noverlap = nperseg // 2

        f, Pxx = welch(
            x, fs=float(self.fs),
            nperseg=nperseg, noverlap=noverlap,
            window="hann", detrend="constant",
            scaling="density"
        )
        Pxx = np.maximum(Pxx, 1e-12)
        return f, Pxx

    def _sos_filter_stream(self, sos, x, key):
        x = np.asarray(x, dtype=float)
        if x.size == 0:
            return x
        if key is None:
            key = ('sos_hash', sos.shape, float(sos.ravel()[0]) if sos.size > 0 else 0.0)

        zi = self._sos_states.get(key)
        if zi is None:
            try:
                zi0 = sp_signal.sosfilt_zi(sos)
                zi = zi0 * float(x[0])
            except Exception:
                zi = sp_signal.sosfilt_zi(sos) * 0.0

        try:
            y, zf = sp_signal.sosfilt(sos, x, zi=zi)
        except Exception as e:
            # fallback: try single-call filtering without zi
            y = sp_signal.sosfilt(sos, x)
            zf = None

        if zf is not None:
            self._sos_states[key] = zf
        return y

    @staticmethod
    def smooth(signal_arr, k=5):
        n = len(signal_arr)
        if n < 5:
            return signal_arr
        w = int(min(k, n if n % 2 == 1 else n - 1))
        if w < 3: w = 3
        if w % 2 == 0:
            w -= 1
            if w < 3: w = 3
        return savgol_filter(signal_arr, window_length=w, polyorder=2)

    @staticmethod
    def assess_signal_quality(segment, fs):
        try:
            from numpy import trapezoid as _trap
        except Exception:
            _trap = np.trapz
        sig = np.asarray(segment, dtype=float)
        n = len(sig)
        if n < 4:
            return {"score": 0, "reason": "segment too short"}
        f, Pxx = welch(sig - np.mean(sig), fs=fs, nperseg=min(1024, n))
        total_power = _trap(Pxx, f)

        def band_energy(lo, hi):
            mask = (f >= lo) & (f <= hi)
            return float(_trap(Pxx[mask], f[mask])) if np.any(mask) else 0.0

        low_energy = band_energy(0.0, 0.5)
        ecg_band_energy = band_energy(0.5, 40.0)
        noise_energy = total_power - ecg_band_energy
        snr = ecg_band_energy / (noise_energy + 1e-12)
        snr_score = np.clip((np.log10(snr + 1) / np.log10(21)), 0, 1)
        baseline_score = 1.0 - np.clip((low_energy / (ecg_band_energy + low_energy + 1e-12)) * 2.0, 0, 1)
        pl_energy = band_energy(49, 51) + band_energy(59, 61)
        powerline_score = 1.0 - np.clip((pl_energy / (total_power + 1e-12)) * 100.0, 0, 1)
        diffs = np.abs(np.ediff1d(sig))
        flat_ratio = np.sum(diffs < (np.std(sig) * 1e-3)) / (len(diffs) + 1e-12)
        flat_score = 1.0 - np.clip(flat_ratio * 10.0, 0, 1)
        peaks, _ = find_peaks(np.abs(sig - np.mean(sig)), height=np.std(sig), distance=int(0.25 * fs))
        rr_consistency = 0.0
        if len(peaks) > 2:
            rr = np.diff(peaks) / fs
            if len(rr) > 3:
                rr_consistency = 1.0 - np.clip(np.std(rr) / (np.mean(rr) + 1e-12), 0, 1)
        tpl_corr = 0.0
        if len(peaks) > 3:
            beats = []
            w = int(0.2 * fs)
            for p in peaks:
                s, e = max(0, p - w // 2), min(len(sig), p + w // 2)
                beat = sig[s:e]
                if len(beat) == w:
                    beats.append(beat)
            if len(beats) >= 3:
                beats = np.array(beats)
                template = np.median(beats, axis=0)
                cors = [np.corrcoef(template, b)[0, 1] for b in beats]
                tpl_corr = np.nanmean([c for c in cors if not np.isnan(c)])
                tpl_corr = (tpl_corr + 1) / 2
        try:
            b, a = signal.butter(2, [0.5 / (fs / 2), 40 / (fs / 2)], btype='band')
            bp = signal.filtfilt(b, a, sig)
            qrs_score = float(np.clip(np.std(bp) / (np.std(sig) + 1e-12), 0, 1))
        except Exception:
            qrs_score = 0.0
        psd_norm = Pxx / (np.sum(Pxx) + 1e-12)
        spec_ent = -np.sum(psd_norm * np.log2(psd_norm + 1e-12)) / np.log2(len(psd_norm))
        spec_ent_score = 1.0 - np.clip(spec_ent, 0, 1)
        skew_val = skew(sig, nan_policy="omit")
        kurt_val = kurtosis(sig, fisher=False, nan_policy="omit")
        mu = np.mean(sig)
        denom = np.sum((sig - mu) ** 2)
        autocorr1 = np.sum((sig[:-1] - mu) * (sig[1:] - mu)) / (denom + 1e-12)
        features = np.array([[snr_score, baseline_score, powerline_score, rr_consistency, tpl_corr, qrs_score,
                              spec_ent_score, autocorr1]])
        clf = LogisticRegression()
        X_train = np.array([[0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9], [0.2, 0.2, 0.2, 0.1, 0.1, 0.2, 0.1, 0.1]])
        y_train = np.array([1, 0])
        clf.fit(X_train, y_train)
        pred = float(clf.predict_proba(features)[0, 1])
        weights = {"snr": 0.25, "rr": 0.20, "tpl": 0.15, "baseline": 0.10, "powerline": 0.10, "qrs": 0.10, "spec": 0.05,
                   "autocorr": 0.03, "skew_kurt": 0.02}
        combined = (snr_score * weights["snr"] + rr_consistency * weights["rr"] + tpl_corr * weights[
            "tpl"] + baseline_score * weights["baseline"] + powerline_score * weights["powerline"] + qrs_score *
                    weights["qrs"] + spec_ent_score * weights["spec"] + autocorr1 * weights["autocorr"] + (
                                (abs(skew_val) + abs(kurt_val - 3)) / 10.0) * weights["skew_kurt"])
        combined = combined * 0.7 + pred * 0.3
        return {"score": float(np.round(combined * 100, 1))}

    #HRV :
    def get_nn_intervals(self, window_seconds: float = None,
                         min_rr_ms: float = 300.0, max_rr_ms: float = 2000.0,
                         reject_pct_change: float = 0.25):
        times = np.asarray(self.all_rr_times, dtype=float)
        if times.size < 2:
            return np.array([]), 0.0

        if window_seconds is not None:
            min_t = times[-1] - float(window_seconds)
            times = times[times >= min_t]
        if times.size < 2:
            return np.array([]), 0.0

        times_used = times.copy()
        rr_ms = np.diff(times_used) * 1000.0
        rr_ms = rr_ms.astype(float)

        phys_mask = (rr_ms >= min_rr_ms) & (rr_ms <= max_rr_ms)
        rr_ms[~phys_mask] = np.nan

        med = np.nanmedian(rr_ms)
        if np.isnan(med):
            raw_med = np.median(np.diff(times_used) * 1000.0)
            med = float(raw_med) if not np.isnan(raw_med) else np.nan
        if np.isnan(med):
            return np.array([]), 0.0

        bad = np.where(np.abs(rr_ms - med) / (med + 1e-12) > reject_pct_change)[0]
        rr_ms[bad] = np.nan

        idx = np.arange(rr_ms.size)
        good = idx[~np.isnan(rr_ms)]
        if good.size >= 2 and np.any(np.isnan(rr_ms)):
            f = interp1d(good, rr_ms[good], kind='linear', fill_value='extrapolate', bounds_error=False)
            nan_idx = idx[np.isnan(rr_ms)]
            rr_ms[nan_idx] = f(nan_idx)

        rr_ms = rr_ms[np.isfinite(rr_ms)]
        actual_window_s = (times_used[-1] - times_used[0]) if times_used.size > 1 else (
            np.sum(rr_ms) / 1000.0 if rr_ms.size else 0.0)
        return rr_ms, float(actual_window_s)

    def time_domain_hrv(self, rr_ms: np.ndarray):
        rr_ms = np.asarray(rr_ms, dtype=float).reshape(-1)
        out = {}
        n = rr_ms.size
        if n < 1:
            return out
        mean_nn = float(np.mean(rr_ms)) if n > 0 else np.nan
        sdnn = float(np.std(rr_ms, ddof=1)) if n > 1 else 0.0
        diffs = np.diff(rr_ms)
        rmssd = float(np.sqrt(np.mean(diffs ** 2))) if diffs.size > 0 else 0.0
        pnn50 = float(100.0 * np.sum(np.abs(diffs) > 50.0) / (len(diffs) if len(diffs) > 0 else 1))
        out["meanNN_ms"] = round(mean_nn, 3)
        out["SDNN_ms"] = round(sdnn, 3)
        out["RMSSD_ms"] = round(rmssd, 3)
        out["pNN50_pct"] = round(pnn50, 3)
        return out

    def poincare(self, rr_ms: np.ndarray):
        if rr_ms is None or len(rr_ms) < 3:
            return {}
        diffs = np.diff(rr_ms)
        var_diff = np.var(diffs, ddof=1)
        var_rr = np.var(rr_ms, ddof=1)
        sd1 = np.sqrt(0.5 * var_diff)
        sd2 = np.sqrt(2.0 * var_rr - 0.5 * var_diff) if (2.0 * var_rr - 0.5 * var_diff) > 0 else 0.0
        return {"SD1_ms": round(float(sd1), 3), "SD2_ms": round(float(sd2), 3)}

    def freq_domain_hrv(self, rr_ms: np.ndarray):
        rr_ms = np.asarray(rr_ms, dtype=float).reshape(-1)
        if rr_ms.size < 3:
            return {}

        t = np.cumsum(rr_ms) / 1000.0
        duration = float(t[-1]) if t.size > 0 else 0.0
        if duration <= 0:
            return {}

        freqs = np.linspace(0.0033, 0.5, 2048)
        ang_freqs = 2 * np.pi * freqs
        y = rr_ms - np.mean(rr_ms)
        pgram = lombscargle(t, y, ang_freqs, precenter=True, normalize=True)

        def band_power(flo, fhi):
            mask = (freqs >= flo) & (freqs < fhi)
            if not np.any(mask):
                return 0.0
            return float(trapezoid(pgram[mask], freqs[mask]))

        vlf = band_power(0.0033, 0.04)
        lf = band_power(0.04, 0.15)
        hf = band_power(0.15, 0.40)

        total = float(vlf + lf + hf)
        lf_hf = float(lf / hf) if hf > 0.0 else float('nan')
        denom = max(total - vlf, 1e-12)
        lf_norm = 100.0 * lf / denom
        hf_norm = 100.0 * hf / denom

        return {
            "VLF": float(vlf),
            "LF": float(lf),
            "HF": float(hf),
            "LF_norm_pct": round(float(lf_norm), 3),
            "HF_norm_pct": round(float(hf_norm), 3),
            "LF_HF": round(float(lf_hf), 4) if not np.isnan(lf_hf) else float('nan'),
            "method": "lomb",
            "duration_s": round(float(duration), 3)
        }

    def sample_entropy(self, x, m=2, r=None):
        x = np.asarray(x, dtype=float)
        N = len(x)
        if N <= m + 1:
            return np.nan
        if r is None:
            r = 0.2 * np.std(x, ddof=1)

        def _phi(m):
            C = 0
            M = N - m + 1
            for i in range(M):
                xi = x[i:i + m]
                for j in range(i + 1, M):
                    xj = x[j:j + m]
                    if np.max(np.abs(xi - xj)) <= r:
                        C += 1
            return C

        A = _phi(m + 1)
        B = _phi(m)
        if B == 0:
            return np.nan
        return -np.log(A / B) if A > 0 else np.nan

    def multi_scale_entropy(self, x, scales=20, m=2, r=None):
        x = np.asarray(x, dtype=float)
        if len(x) < m + 1:
            return np.full(scales, np.nan)
        if r is None:
            r = 0.15 * np.std(x, ddof=1)

        mse = []
        for s in range(1, scales + 1):
            if len(x) // s < m + 1:
                mse.append(np.nan)
                continue
            y = np.mean(x[:len(x) - len(x) % s].reshape(-1, s), axis=1)
            mse.append(self.sample_entropy(y, m=m, r=r))
        return np.array(mse)

    def detrended_fluctuation_analysis(self, x):
        x = np.asarray(x, dtype=float)
        N = len(x)
        if N < 10:
            return np.nan, np.nan

        y = np.cumsum(x - np.mean(x))
        scales = np.unique(np.floor(np.logspace(np.log10(4), np.log10(N // 4), num=20))).astype(int)
        F = []
        for s in scales:
            segments = N // s
            rms = []
            for i in range(segments):
                seg = y[i * s:(i + 1) * s]
                t = np.arange(s)
                coeffs = np.polyfit(t, seg, 1)
                trend = np.polyval(coeffs, t)
                rms.append(np.sqrt(np.mean((seg - trend) ** 2)))
            F.append(np.mean(rms))

        log_scales = np.log(scales)
        log_F = np.log(F)
        alpha, _ = np.polyfit(log_scales, log_F, 1)
        return float(alpha), float(np.mean(F))

    def compute_hrv(self, window_seconds: float = None, compute_freq_domain: bool = True):
        rr_pack = self.get_nn_intervals(window_seconds=window_seconds)
        if isinstance(rr_pack, tuple) and len(rr_pack) == 2:
            rr_ms, actual_window_s = rr_pack
        else:
            rr_ms = np.asarray(rr_pack, dtype=float).reshape(-1)
            actual_window_s = float(np.sum(rr_ms) / 1000.0) if rr_ms.size else 0.0

        rr_ms = np.asarray(rr_ms, dtype=float).reshape(-1)
        out = {"rr_ms": rr_ms, "beats_used": int(rr_ms.size), "actual_window_s": float(actual_window_s)}

        if rr_ms.size < 2:
            out["time_domain"] = {}
            out["freq_domain"] = {}
            out["sample_entropy"] = float('nan')
            out["multi_scale_entropy"] = []
            out["dfa_alpha"] = float('nan')
            out["dfa_fluct"] = float('nan')
            return out

        out["time_domain"] = self.time_domain_hrv(rr_ms)
        fd = {}
        beats_needed = 5
        if compute_freq_domain and rr_ms.size >= beats_needed:
            try:
                fd = self.freq_domain_hrv(rr_ms)
            except Exception as e:
                print("Frequency domain HRV error:", e)
                fd = {}
        else:
            fd = {}

        out["freq_domain"] = fd

        try:
            out["sample_entropy"] = float(self.sample_entropy(rr_ms)) if rr_ms.size >= 10 else float('nan')
        except Exception:
            out["sample_entropy"] = float('nan')

        try:
            out["multi_scale_entropy"] = self.multi_scale_entropy(rr_ms, scales=20).tolist() if rr_ms.size >= 10 else []
        except Exception:
            out["multi_scale_entropy"] = []

        try:
            a, f = self.detrended_fluctuation_analysis(rr_ms) if rr_ms.size >= 10 else (float('nan'), float('nan'))
            out["dfa_alpha"] = a
            out["dfa_fluct"] = f
        except Exception:
            out["dfa_alpha"] = float('nan')
            out["dfa_fluct"] = float('nan')

        return out

    #Pipeline management
    def add_filter(self, name: str, func):
        if name in self.filter_names:
            return False
        if not callable(func):
            raise ValueError("Filter function must be callable (e.g., self.low/self.high/self.band)")
        self.filters_pipeline.append((name, func))
        self.filter_names.add(name)
        return True

    def remove_filter(self, name: str):
        if name not in self.filter_names:
            return False
        self.filters_pipeline = [(n, f) for n, f in self.filters_pipeline if n != name]
        self.filter_names.remove(name)
        remaining = [n.lower() for n, _ in self.filters_pipeline]
        if not any("low" in n for n in remaining):
            keys = [k for k in list(self._filter_cache.keys()) if k[0] == 'low']
            for k in keys:
                del self._filter_cache[k]
        if not any("high" in n for n in remaining):
            keys = [k for k in list(self._filter_cache.keys()) if k[0] == 'high']
            for k in keys:
                del self._filter_cache[k]
        if not any("band" in n for n in remaining):
            keys = [k for k in list(self._filter_cache.keys()) if k[0] == 'band']
            for k in keys:
                del self._filter_cache[k]
        self._sos_states = {k: v for k, v in self._sos_states.items() if k[0] != name}
        return True

    def get_active_filters(self):
        return [n for n, _ in self.filters_pipeline]

    def reset_signal(self):
        self._filter_cache.clear()
        self.filters_pipeline.clear()
        self.filter_names.clear()
        self.current_signal = self.original_ecg.copy()
        self.all_rr_times.clear()
        self.last_segment = None
        self._sos_states.clear()

    #Hr detection
    def detect_heart_rate(self, segment, start_index):
        fs = float(self.fs)
        x = np.asarray(segment, dtype=float)
        if x.size < max(3, int(1.0 * fs)):
            return 0.0
        try:
            bp = self.band(x, fs, 5.0, 18.0, order=3, zero_phase=False, streaming=True)
        except Exception:
            bp = x
        d = np.diff(bp)
        sq = d * d
        win = max(1, int(0.150 * fs))
        ma = np.convolve(sq, np.ones(win) / win, mode="same")
        init = min(len(ma), int(2 * fs))
        mu = float(np.mean(ma[:init])) if init else float(np.mean(ma))
        sd = float(np.std(ma[:init], ddof=1)) if init > 1 else 0.0
        if getattr(self, "signal_level", 0.0) == 0.0 and getattr(self, "noise_level", 0.0) == 0.0 and getattr(self,
                                                                                                              "thresh_level",
                                                                                                              0.0) == 0.0:
            self.signal_level = mu
            self.noise_level = sd
            self.thresh_level = mu + 0.5 * sd
        rr_hist = np.diff(self.all_rr_times[-8:]) if len(self.all_rr_times) >= 9 else np.array([])
        mean_rr = float(np.median(rr_hist)) if rr_hist.size else 0.0
        refr = int(max(0.20, min(0.40, 0.6 * mean_rr if mean_rr > 0 else 0.20)) * fs)
        peaks0 = []
        last = -refr
        for i in range(1, len(ma) - 1):
            is_peak = (ma[i] > self.thresh_level) and (ma[i] > ma[i - 1]) and (ma[i] > ma[i + 1])
            if is_peak and (i - last > refr):
                peaks0.append(i)
                last = i
                self.signal_level = 0.125 * ma[i] + 0.875 * self.signal_level
            else:
                self.noise_level = 0.125 * ma[i] + 0.875 * self.noise_level
            self.thresh_level = self.noise_level + 0.25 * (self.signal_level - self.noise_level)
        Lmax = min(4, pywt.dwt_max_level(len(bp), pywt.Wavelet('db4').dec_len))
        coeffs = pywt.wavedec(bp, 'db4', level=max(1, Lmax))
        dets = []
        for li in range(1, min(4, len(coeffs))):
            dets.append(np.abs(coeffs[li]))
        if len(dets) == 0:
            dets = [np.abs(bp)]
        wlens = []
        for li in range(len(dets)):
            wlens.append(int(round(len(bp) / len(dets[li]))))
        mad_thr = []
        for det in dets:
            med = np.median(det)
            mad = np.median(np.abs(det - med)) + 1e-12
            mad_thr.append(med + 3.5 * mad)
        verified = []
        for p in peaks0:
            ok = False
            for det, wlen, thr in zip(dets, wlens, mad_thr):
                idx = int(round(p / max(1, wlen)))
                s = max(0, idx - int(0.06 * fs / max(1, wlen)))
                e = min(len(det), idx + int(0.06 * fs / max(1, wlen)))
                if e > s and np.max(det[s:e]) > thr:
                    ok = True
                    break
            if ok:
                verified.append(p)
        templ_ok = []
        if not hasattr(self, "_qrs_template"):
            self._qrs_template = None
        for p in verified:
            w = int(0.180 * fs)
            s = max(0, p - w // 2)
            e = min(len(bp), p + w // 2)
            seg = bp[s:e]
            take = True
            if self._qrs_template is not None and len(seg) == len(self._qrs_template):
                a = seg - np.mean(seg)
                b = self._qrs_template - np.mean(self._qrs_template)
                na = np.linalg.norm(a) + 1e-12
                nb = np.linalg.norm(b) + 1e-12
                corr = float(np.dot(a, b) / (na * nb))
                take = corr >= 0.15
            if take:
                templ_ok.append(p)
        refined_times = []
        for p in templ_ok:
            try:
                t_ref = (start_index / fs) + self._refine_peak_time(bp, p, fs)
                refined_times.append(t_ref)
            except Exception:
                refined_times.append((start_index + p) / fs)
        if refined_times:
            w = int(0.180 * fs)
            beats = []
            for p in templ_ok:
                s = max(0, p - w // 2)
                e = min(len(bp), p + w // 2)
                if e - s == w:
                    beats.append(bp[s:e])
            if len(beats) >= 3:
                self._qrs_template = np.median(np.stack(beats, axis=0), axis=0)
            elif self._qrs_template is None and len(beats) >= 1:
                self._qrs_template = beats[0]
        new_times = []
        for t in refined_times:
            if not self.all_rr_times:
                new_times.append(t)
            else:
                rr_new = t - self.all_rr_times[-1]
                if rr_new >= 0.20:
                    if len(self.all_rr_times) >= 8:
                        last_rr = np.diff(self.all_rr_times[-8:])
                        mrr = np.median(last_rr) if last_rr.size else 0.0
                        if mrr > 0.0:
                            if 0.5 * mrr <= rr_new <= 1.8 * mrr:
                                new_times.append(t)
                        else:
                            new_times.append(t)
                    else:
                        new_times.append(t)
        for t in new_times:
            self.all_rr_times.append(t)
        if self.all_rr_times:
            t0 = self.all_rr_times[-1] - float(self.history_seconds)
            self.all_rr_times = [t for t in self.all_rr_times if t >= t0]
        if len(self.all_rr_times) >= 3:
            rr = np.diff(self.all_rr_times[-min(12, len(self.all_rr_times)):])
            if rr.size:
                rr_med = float(np.median(rr))
                if rr_med > 0:
                    return 60.0 / rr_med
        return 0.0

    def _refine_peak_time(self, qrs_sig, peak_idx, fs, search_ms=80):
        w = max(3, int((search_ms / 1000.0) * fs))
        s = max(1, peak_idx - w)
        e = min(len(qrs_sig) - 2, peak_idx + w)
        loc = s + np.argmax(np.abs(qrs_sig[s:e]))
        y1, y2, y3 = qrs_sig[loc - 1], qrs_sig[loc], qrs_sig[loc + 1]
        denom = (y1 - 2 * y2 + y3)
        if denom == 0:
            frac = 0.0
        else:
            frac = 0.5 * (y1 - y3) / denom  # إزاحة تحت-العينة
        return (loc + frac) / fs  # زمن بالثواني

    def _moving_average(self, sig, win_s):
        L = max(1, int(round(win_s * self.fs)))
        return np.convolve(sig, np.ones(L, dtype=float) / float(L), mode='same')


def start_main_app(path):
    try:
        processor = SignalProcessor(path)
        app = ECGApp(processor)
        app.run()
    except Exception as e:
        messagebox.showerror("Error", f"Failed to start app: {e}")


class ECGApp:
    def __init__(self, processor):
        self.processor = processor
        self.ani2 = None
        self.filters_win = None
        self.filter_controls = {}
        #HRV update control
        self.hrv_update_interval = 2.0
        self._last_hrv_update = 0.0

        # GUI settings::
        self.root = tk.Tk()
        self.root.title("Live ECG Processor")
        self.root.geometry("1300x800")

        self.frame_left = tk.Frame(self.root, bg="#2B2B2B")
        self.frame_left.pack(side=tk.LEFT, padx=10, pady=10)

        self.frame_plot = tk.Frame(self.root, bg="#2B2B2B")
        self.frame_plot.pack(side=tk.RIGHT)

        # plot setting
        self.fig, self.ax = plt.subplots(figsize=(10, 4))
        self.x = np.arange(self.processor.window_size)
        self.y = self.processor.current_signal[:self.processor.window_size]
        self.line, = self.ax.plot(self.x, self.y, linewidth=1, color="cyan")
        self.ax.set_facecolor("#1E1E1E")
        self.fig.patch.set_facecolor("#1E1E1E")
        self.ax.set_title("Live ECG Signal", color="white")
        self.ax.set_xlabel("Time (s)", fontsize=12, color="white")
        self.ax.set_ylabel("Amplitude", fontsize=12, color="white")
        self.ax.tick_params(axis="x", colors="white")
        self.ax.tick_params(axis="y", colors="white")
        self.ax.grid(True, which="both", linestyle="--", alpha=0.5)

        plt.tight_layout()

        self.canvas = FigureCanvasTkAgg(self.fig, master=self.frame_plot)
        self.canvas_widget = self.canvas.get_tk_widget()
        self.canvas_widget.pack()

        self.calc_dots = 0
        self.calc_running = False

        def animate_calculating():
            if not self.calc_running:
                return
            dots = "." * (self.calc_dots % 4)
            self.hr_label.config(text=f"HR: Calculating{dots}")
            self.calc_dots += 1
            self.root.after(500, animate_calculating)


        # animation
        def update(frame):
            if not self.processor.is_running:
                return self.line,

            if self.processor.index + self.processor.window_size >= len(self.processor.original_ecg):
                self.processor.index = 0

            self.processor.segment = self.processor.original_ecg[
                                     self.processor.index:self.processor.index + self.processor.window_size].copy()

            for name, f in self.processor.filters_pipeline:
                self.processor.segment = f(self.processor.segment)

            self.processor.last_segment = self.processor.segment.copy()
            start_time = self.processor.index / self.processor.fs
            end_time = (self.processor.index + self.processor.window_size) / self.processor.fs
            x = np.linspace(start_time, end_time, self.processor.window_size, endpoint=False)
            self.line.set_xdata(x)
            self.line.set_ydata(self.processor.segment)
            self.ax.set_xlim(start_time, end_time)
            self.ax.set_ylim(np.min(self.processor.segment) - 0.5, np.max(self.processor.segment) + 0.5)
            self.ax.xaxis.set_major_formatter(ticker.FormatStrFormatter('%.2f'))

            hr = float(self.processor.detect_heart_rate(self.processor.segment, self.processor.index))
            if hr > 0:
                self.hr_label.config(text=f"HR: {hr:.1f} BPM")
                self.calc_running = False
            else:
                if not self.calc_running:
                    self.calc_running = True
                    animate_calculating()

            self.processor.index += self.processor.step

            try:
                sqa = SignalProcessor.assess_signal_quality(self.processor.segment, self.processor.fs)
                new_score = float(sqa.get("score", 0.0))
            except Exception:
                new_score = 0.0
            alpha = 0.20
            if self._sqa_smooth is None:
                self._sqa_smooth = new_score
            else:
                self._sqa_smooth = (1 - alpha) * self._sqa_smooth + alpha * new_score

            disp = round(self._sqa_smooth, 1)
            bg = self._score_to_color(disp)
            fg = self._best_text_color(bg)
            self.sqa_label.config(text=f"Quality: {disp:.1f} %  ({self._quality_band(disp)})", bg=bg, fg=fg)

            try:
                now_ts = time.time()
                if (now_ts - self._last_hrv_update) >= self.hrv_update_interval:
                    self._last_hrv_update = now_ts
                    beats_needed = getattr(self.processor, "hrv_freq_min_beats", 5)
                    current_beats = len(getattr(self.processor, "all_rr_times", []))
                    compute_freq = True if current_beats >= int(beats_needed) else False
                    hrv = self.processor.compute_hrv(window_seconds=None, compute_freq_domain=compute_freq)

                    rr = np.asarray(hrv.get("rr_ms", []), dtype=float).reshape(-1)
                    beats_used = int(hrv.get("beats_used", rr.size))
                    actual_window_s = float(hrv.get("actual_window_s", (np.sum(rr) / 1000.0) if rr.size else 0.0))

                    td = hrv.get("time_domain", {}) or {}
                    fd = hrv.get("freq_domain", {}) or {}

                    rmssd = td.get("RMSSD_ms", "--")
                    sdnn = td.get("SDNN_ms", "--")
                    lf_hf = fd.get("LF_HF", "--") if fd else "--"

                    status = ""
                    if not compute_freq:
                        status = f" (collecting data: {current_beats}/{beats_needed} beats)"
                        if hasattr(self, "data_status_label") and getattr(self, "data_status_label") is not None:
                            self.data_status_label.config(
                                text=f"Collecting RR data: {current_beats}/{beats_needed} beats", fg="#ffcc00")
                    else:
                        if hasattr(self, "data_status_label") and getattr(self, "data_status_label") is not None:
                            self.data_status_label.config(text="HRV analysis active", fg="lime")

                    txt = f"HRV: RMSSD {rmssd} ms | SDNN {sdnn} ms | LF/HF {lf_hf}{status}"

                    if hasattr(self, "hrv_summary_label") and getattr(self, "hrv_summary_label") is not None:
                        self.hrv_summary_label.config(text=txt)
                    if hasattr(self, "hrv_under_plot_label") and getattr(self, "hrv_under_plot_label") is not None:
                        self.hrv_under_plot_label.config(text=txt)
            except Exception as e:
                print("HRV update error:", e)

            return self.line,

        self.ani = animation.FuncAnimation(self.fig, update, interval=65, blit=False, cache_frame_data=False)

        # Frame
        self.controls_frame = tk.Frame(self.frame_left, bg="#1E1E1E", padx=10, pady=10)
        self.controls_frame.pack(fill="both", expand=True)

        self.style = ttk.Style()
        self.style.configure("TButton",
                        font=("Segoe UI", 11, "bold"),
                        padding=8,
                        background="#2D89EF",
                        foreground="black")
        self.style.map("TButton",
                  foreground=[("active", "black")],
                  background=[("active", "#1BA1E2")])

        self.style.configure("TLabel",
                        font=("Segoe UI", 10),
                        background="#1E1E1E",
                        foreground="white")

        self.style.configure("TEntry",
                        font=("Segoe UI", 10),
                        fieldbackground="white",
                        foreground="black")

        self.style.configure("Filter.TCombobox",
                             font=("Segoe UI", 11),
                             foreground="white")

        self.style.configure("Filter.TCombobox",
                             fieldbackground="#1E1E1E",
                             background="#1E1E1E",
                             foreground="white",
                             padding=4,
                             arrowcolor="white")

        self.style.configure("TCombobox",
                             fieldbackground="#1E1E1E",
                             background="#1E1E1E",
                             foreground="white")

        self.style.map("Filter.TCombobox",
                       fieldbackground=[("readonly", "#1E1E1E")],
                       background=[("readonly", "#1E1E1E")],
                       foreground=[("readonly", "white")])

        # Quality label
        self.sqa_label = tk.Label(
            self.frame_plot,
            text="Quality: -- %",
            font=("Segoe UI", 12, "bold"),
            fg="white",
            bg="#e74c3c",
            padx=8, pady=4
        )
        self.sqa_label.pack(pady=(0, 6))
        self._sqa_smooth = None

        # Grids
        self.controls_frame.grid_columnconfigure(0, weight=1)
        self.controls_frame.grid_columnconfigure(1, weight=1)

        # Reset
        ttk.Button(self.controls_frame, text="Reset (Raw)", command=self.reset_signal) \
            .grid(row=0, column=0, columnspan=2, pady=5, sticky="ew")

        # Low Pass
        ttk.Label(self.controls_frame, text="Low Pass Cutoff:") \
            .grid(row=1, column=0, sticky="w")
        self.entry_low_cutoff = ttk.Entry(self.controls_frame)
        self.entry_low_cutoff.insert(0, "40")
        self.entry_low_cutoff.grid(row=1, column=1, sticky="ew")
        ttk.Button(self.controls_frame, text="Apply Low Pass", command=self.apply_lowpass) \
            .grid(row=2, column=0, columnspan=2, pady=5, sticky="ew")

        # High Pass
        ttk.Label(self.controls_frame, text="High Pass Cutoff:") \
            .grid(row=3, column=0, sticky="w")
        self.entry_high_cutoff = ttk.Entry(self.controls_frame)
        self.entry_high_cutoff.insert(0, "0.5")
        self.entry_high_cutoff.grid(row=3, column=1, sticky="ew")
        ttk.Button(self.controls_frame, text="Apply High Pass", command=self.apply_highpass) \
            .grid(row=4, column=0, columnspan=2, pady=5, sticky="ew")

        # Band Pass
        ttk.Label(self.controls_frame, text="Band Pass Low:") \
            .grid(row=5, column=0, sticky="w")
        self.entry_band_low = ttk.Entry(self.controls_frame)
        self.entry_band_low.insert(0, "0.5")
        self.entry_band_low.grid(row=5, column=1, sticky="ew")

        ttk.Label(self.controls_frame, text="Band Pass High:") \
            .grid(row=6, column=0, sticky="w")
        self.entry_band_high = ttk.Entry(self.controls_frame)
        self.entry_band_high.insert(0, "40")
        self.entry_band_high.grid(row=6, column=1, sticky="ew")

        ttk.Button(self.controls_frame, text="Apply Band Pass", command=self.apply_bandpass) \
            .grid(row=7, column=0, columnspan=2, pady=5, sticky="ew")

        # Smooth
        ttk.Label(self.controls_frame, text="Smooth Kernel Size:") \
            .grid(row=8, column=0, sticky="w")
        self.entry_smooth = ttk.Entry(self.controls_frame)
        self.entry_smooth.insert(0, "5")
        self.entry_smooth.grid(row=8, column=1, sticky="ew")
        ttk.Button(self.controls_frame, text="Apply Smoothing", command=self.apply_smooth) \
            .grid(row=9, column=0, columnspan=2, pady=5, sticky="ew")

        # Zero-phase
        self.zero_phase_enabled = False

        self.btn_zero = tk.Button(
            self.controls_frame,
            text="Zero-phase: OFF",
            command=self.toggle_zero_phase,
            font=("Segoe UI", 11, "bold"),
            fg="white",
            bg="#e74c3c",
            activebackground="#c0392b",
            relief="raised",
            padx=8,
            pady=6
        )
        self.btn_zero.grid(row=10, column=0, columnspan=2, sticky="ew", pady=8, padx=4)

        # Fourier
        ttk.Button(self.controls_frame, text="Show Fourier Transform", command=self.show_fourier) \
            .grid(row=11, column=0, columnspan=2, pady=10, sticky="ew")

        # Manage Filters
        ttk.Button(self.controls_frame, text="Manage Filters", command=self.open_filters_manager) \
            .grid(row=12, column=0, columnspan=2, pady=5, sticky="ew")


        #HRV buttons
        ttk.Button(self.controls_frame, text="Show HRV", command=self.show_hrv_window) \
            .grid(row=14, column=0, sticky="ew", padx=(0,4), pady=6)
        ttk.Button(self.controls_frame, text="Export HRV CSV", command=self.export_hrv_csv_dialog) \
            .grid(row=14, column=1, sticky="ew", padx=(4,0), pady=6)

        self.data_status_label = tk.Label(self.root, text="Collecting RR data for HRV...", font=("Segoe UI", 9),
                                          fg="#ffcc00", bg="#1E1E1E")
        self.data_status_label.pack(side="bottom", pady=(2, 5))

        # Pause/Resume
        self.btn_toggle = ttk.Button(self.controls_frame, text="Pause", command=self.toggle_run)
        self.btn_toggle.grid(row=13, column=0, columnspan=2, pady=20, sticky="ew")

        # Heart rate
        self.hr_label = ttk.Label(self.frame_plot, text="HR: -- BPM", font=("Segoe UI", 14, "bold"), foreground="lime",
                             background="#1E1E1E")
        self.hr_label.pack(pady=5)

        self.root.configure(bg="#2B2B2B")

        # HRV
        self.hrv_summary_label = ttk.Label(self.frame_plot, text="HRV: RMSSD -- ms | SDNN -- ms | LF/HF --",
                                           font=("Segoe UI", 10), foreground="white", background="#1E1E1E")
        self.hrv_summary_label.pack(pady=2)

    def update_sqa_label(self):
        """Recompute SQA and update the label immediately"""
        try:
            seg = self.processor.last_segment if self.processor.last_segment is not None else self.processor.segment
            if seg is None or len(seg) < 4:
                return
            sqa = SignalProcessor.assess_signal_quality(seg, self.processor.fs)
            new_score = float(sqa.get("score", 0.0))
            self._sqa_smooth = new_score if self._sqa_smooth is None else 0.5 * self._sqa_smooth + 0.5 * new_score
            disp = round(self._sqa_smooth, 1)
            bg = self._score_to_color(disp)
            fg = self._best_text_color(bg)
            self.sqa_label.config(
                text=f"Quality: {disp:.1f} %  ({self._quality_band(disp)})",
                bg=bg, fg=fg
            )
        except Exception:
            pass

    def show_fourier(self):

        fourier_window = tk.Toplevel(self.root)
        fourier_window.title("Frequency Spectrum (Welch PSD)")
        fourier_window.geometry("900x520")

        fig2 = plt.Figure(figsize=(9, 5.2), dpi=100)
        ax2 = fig2.add_subplot(111)

        line_freq, = ax2.plot([], [], linewidth=1, color="cyan")
        ax2.set_xlim(0, 100)
        ax2.set_ylim(-120, 0)
        ax2.set_xlabel("Frequency (Hz)")
        ax2.set_ylabel("Power/Frequency (dB/Hz)")
        ax2.set_title("Power Spectral Density")

        ax2.set_facecolor("#1E1E1E");
        fig2.patch.set_facecolor("#1E1E1E")
        ax2.grid(True, which="both", linestyle="--", alpha=0.5)
        ax2.tick_params(axis="x", colors="white")
        ax2.tick_params(axis="y", colors="white")
        ax2.title.set_color("white");
        ax2.xaxis.label.set_color("white");
        ax2.yaxis.label.set_color("white")

        canvas2 = FigureCanvasTkAgg(fig2, master=fourier_window)
        canvas2.draw()
        canvas2.get_tk_widget().pack(fill="both", expand=True)

        def update_fourier(frame):
            if not self.processor.is_running:
                return line_freq,
            if self.processor.last_segment is not None:
                sig = self.processor.last_segment.copy()
            else:
                sig = self.processor.original_ecg[:self.processor.window_size].copy()
                for name, f in self.processor.filters_pipeline:
                    sig = f(sig)

            sig = sig - np.mean(sig)
            nperseg = min(1024, len(sig))
            f, px = self.processor.compute_spectrum(sig)
            px = 10 * np.log10(px + 1e-12)
            line_freq.set_data(f, px)
            return line_freq,

        self.ani2 = FuncAnimation(fig2, update_fourier, interval=500, blit=True,cache_frame_data=False)
        canvas2.draw()

    def reset_signal(self):
        self.processor.reset_signal()
        if getattr(self, "filter_controls", None):
            self.update_filters_manager_ui()
        self._sqa_smooth = None
        self.sqa_label.config(text="Quality: -- %", bg="#e74c3c", fg="white")

    def toggle_run(self):
        self.processor.is_running = not self.processor.is_running
        self.btn_toggle.config(text="Resume" if not self.processor.is_running else "Pause")
        if self.ani2 is not None:
            if self.processor.is_running:
                self.ani2.event_source.start()
            else:
                self.ani2.event_source.stop()

    def apply_lowpass(self):
        try:
            cutoff = float(self.entry_low_cutoff.get())
        except:
            cutoff = 40.0
        name = f"low_cutoff={cutoff}"
        added = self.processor.add_filter(
            name,
            lambda sig, c=cutoff: self.processor.low(
                sig, self.processor.fs, c,
                zero_phase=self.processor.use_zero_phase
            )
        )
        if not added:
            messagebox.showinfo("Info", f"Filter already active: {name}")
            return
        if hasattr(self, "update_filters_manager_ui"):
            self.update_filters_manager_ui()
        self.update_sqa_label()

    def apply_highpass(self):
        try:
            cutoff = float(self.entry_high_cutoff.get())
        except:
            cutoff = 0.5
        name = f"high_cutoff={cutoff}"
        added = self.processor.add_filter(
            name,
            lambda sig, c=cutoff: self.processor.high(
                sig, self.processor.fs, c,
                zero_phase=self.processor.use_zero_phase
            )
        )
        if not added:
            messagebox.showinfo("Info", f"Filter already active: {name}")
            return
        if hasattr(self, "update_filters_manager_ui"):
            self.update_filters_manager_ui()
        self.update_sqa_label()

    def apply_bandpass(self):
        try:
            lowcut = float(self.entry_band_low.get())
            highcut = float(self.entry_band_high.get())
            if not (0 < lowcut < highcut < self.processor.fs / 2):
                raise ValueError
        except Exception:
            messagebox.showerror("Error", "Band limits invalid. Ensure 0 < low < high < fs/2")
            return
        name = f"band_{lowcut}_{highcut}"
        added = self.processor.add_filter(
            name,
            lambda sig, l=lowcut, h=highcut: self.processor.band(
                sig, self.processor.fs, l, h,
                zero_phase=self.processor.use_zero_phase
            )
        )
        if not added:
            messagebox.showinfo("Info", f"Filter already active: {name}")
            return
        if hasattr(self, "update_filters_manager_ui"):
            self.update_filters_manager_ui()
        self.update_sqa_label()

    def apply_smooth(self):
        try:
            val = int(self.entry_smooth.get())
        except:
            val = 5
        name = f"smooth_k={val}"
        added = self.processor.add_filter(name, lambda sig, k=val: self.processor.smooth(sig, k))
        if not added:
            messagebox.showinfo("Info", f"Filter already active: {name}")
            return
        if hasattr(self, "update_filter_dropdown"):
            self.update_filter_dropdown()
        if hasattr(self, "update_filters_manager_ui"):
            self.update_filters_manager_ui()
        self.update_sqa_label()

    def open_filters_manager(self):
        if hasattr(self, "filters_win") and getattr(self, "filters_win") and self.filters_win.winfo_exists():
            self.filters_win.lift()
            return

        self.filters_win = tk.Toplevel(self.root)
        self.filters_win.title("Filters Manager")
        self.filters_win.geometry("570x255")
        self.filters_win.configure(bg="#1E1E1E")
        self.filters_win.resizable(False, False)
        self.filter_controls = {}

        families = [
            ("low", "Low Pass"),
            ("high", "High Pass"),
            ("band", "Band Pass"),
            ("smooth", "Smoothing")
        ]

        for i, (prefix, label_text) in enumerate(families):
            lbl = tk.Label(self.filters_win, text=label_text + ":", font=("Segoe UI", 11, "bold"),
                           bg="#1E1E1E", fg="white")
            lbl.grid(row=i, column=0, sticky="w", padx=10, pady=8)

            combo_var = tk.StringVar()
            combo = ttk.Combobox(self.filters_win, textvariable=combo_var, state="readonly",
                                 width=34, style="Filter.TCombobox")
            combo.grid(row=i, column=1, sticky="w", padx=(0, 8))

            btn = ttk.Button(self.filters_win, text="Remove",
                             command=lambda p=prefix, c=combo: self._remove_filter_from_manager(p, c))
            btn.grid(row=i, column=2, padx=6)
            btn.state(["disabled"])

            self.filter_controls[prefix] = (combo_var, combo, btn)

        ttk.Button(self.filters_win, text="Close", command=self.filters_win.destroy).grid(
            row=5, column=0, columnspan=3, pady=12
        )
        self.update_filters_manager_ui()

    def update_filters_manager_ui(self):
        if not getattr(self, "filter_controls", None):
            return
        if not hasattr(self, "filters_win") or not getattr(self, "filters_win") or not self.filters_win.winfo_exists():
            return

        try:
            active = self.processor.get_active_filters()
        except Exception:
            active = []

        for family, (combo_var, combo, btn) in self.filter_controls.items():
            try:
                if not combo.winfo_exists():
                    continue
            except tk.TclError:
                continue

            matches = [n for n in active if n.lower().startswith(family)]
            try:
                combo['values'] = matches
                if matches:
                    combo_var.set(matches[0])
                    btn.state(["!disabled"])
                else:
                    combo_var.set("")
                    btn.state(["disabled"])
            except tk.TclError:
                continue

    def _remove_filter_from_manager(self, family_prefix, combo_widget):
        try:
            if combo_widget is None or not combo_widget.winfo_exists():
                messagebox.showinfo("Info", "Combobox not available.")
                return
        except tk.TclError:
            messagebox.showinfo("Info", "Combobox not available.")
            return

        try:
            active = self.processor.get_active_filters()
        except Exception:
            active = []

        matches = [n for n in active if n.lower().startswith(family_prefix)]
        if not matches:
            messagebox.showinfo("Info", f"No active {family_prefix} filter to remove.")
            self.update_filters_manager_ui()
            return

        try:
            selected = combo_widget.get()
        except tk.TclError:
            selected = ""

        to_remove = selected if selected in matches else matches[0]

        removed = self.processor.remove_filter(to_remove)
        if not removed:
            messagebox.showinfo("Info", f"Could not remove filter: {to_remove}")
            self.update_filters_manager_ui()
            return
        self.update_filters_manager_ui()
        try:
            start_idx = self.processor.index
            seg = self.processor.original_ecg[start_idx:start_idx + self.processor.window_size].copy()
            for name, f in self.processor.filters_pipeline:
                seg = f(seg)

            self.processor.last_segment = seg.copy()
            x = np.linspace(start_idx / self.processor.fs,
                            (start_idx + self.processor.window_size) / self.processor.fs,
                            self.processor.window_size, endpoint=False)
            self.line.set_xdata(x)
            self.line.set_ydata(seg)
            self.ax.set_ylim(np.min(seg) - 0.5, np.max(seg) + 0.5)
            self.canvas.draw()
        except Exception:
            pass

        self.update_sqa_label()

    def toggle_zero_phase(self):

        self.zero_phase_enabled = not getattr(self, "zero_phase_enabled", False)
        self.processor.use_zero_phase = self.zero_phase_enabled

        if self.zero_phase_enabled:
            self.btn_zero.config(text="Zero-phase: ON", bg="#2ecc71", activebackground="#27ae60")
        else:
            self.btn_zero.config(text="Zero-phase: OFF", bg="#e74c3c", activebackground="#c0392b")

        try:
            self.processor._sos_states.clear()
        except Exception:
            pass

        try:
            start_idx = self.processor.index
            seg = self.processor.original_ecg[start_idx:start_idx + self.processor.window_size].copy()
            for name, f in self.processor.filters_pipeline:
                seg = f(seg)
            self.processor.last_segment = seg.copy()

            x = np.linspace(start_idx / self.processor.fs,
                            (start_idx + self.processor.window_size) / self.processor.fs,
                            self.processor.window_size, endpoint=False)
            self.line.set_xdata(x)
            self.line.set_ydata(seg)

            try:
                self.ax.set_ylim(np.min(seg) - 0.5, np.max(seg) + 0.5)
            except Exception:
                pass
            self.canvas.draw()
        except Exception as e:
            print("toggle_zero_phase redraw error:", e)

        self.update_sqa_label()

    def _score_to_color(self, score_pct: float) -> str:
        s = max(0.0, min(100.0, float(score_pct)))
        hue = (s / 100.0) * (120.0 / 360.0)
        r, g, b = colorsys.hsv_to_rgb(hue, 0.85, 0.95)
        return f"#{int(r * 255):02x}{int(g * 255):02x}{int(b * 255):02x}"

    def _best_text_color(self, hex_color: str) -> str:
        try:
            r = int(hex_color[1:3], 16) / 255.0
            g = int(hex_color[3:5], 16) / 255.0
            b = int(hex_color[5:7], 16) / 255.0
        except Exception:
            return "white"
        L = 0.2126 * r + 0.7152 * g + 0.0722 * b
        return "black" if L > 0.65 else "white"

    def _quality_band(self, score_pct: float) -> str:
        s = float(score_pct)
        if s >= 80: return "Excellent"
        if s >= 60: return "Good"
        if s >= 40: return "Fair"
        return "Poor"

    def show_hrv_window(self):
        win = tk.Toplevel(self.root)
        win.transient(self.root)
        win.title("HRV Analysis")
        w, h = 1100, 760
        ws = win.winfo_screenwidth()
        hs = win.winfo_screenheight()
        x = int((ws - w) / 2)
        y = int((hs - h) / 2)
        win.geometry(f"{w}x{h}+{x}+{y}")
        win.configure(bg="#1E1E1E")

        title_font = ("Segoe UI", 12, "bold")
        metric_font = ("Segoe UI", 11)
        small_font = ("Segoe UI", 10)

        top_bar = tk.Frame(win, bg="#1E1E1E")
        top_bar.pack(fill="x", padx=12, pady=(10, 6))

        lbl_title = tk.Label(top_bar, text="HRV Live Analysis", font=("Segoe UI", 14, "bold"),
                             fg="white", bg="#1E1E1E")
        lbl_title.pack(side="left", padx=(6, 12))

        metrics_frame = tk.Frame(top_bar, bg="#1E1E1E")
        metrics_frame.pack(side="left", padx=6)

        lbl_rmssd = tk.Label(metrics_frame, text="RMSSD: -- ms", font=metric_font, fg="white", bg="#1E1E1E")
        lbl_rmssd.grid(row=0, column=0, padx=8)
        lbl_sdnn = tk.Label(metrics_frame, text="SDNN: -- ms", font=metric_font, fg="white", bg="#1E1E1E")
        lbl_sdnn.grid(row=0, column=1, padx=8)
        lbl_lfhf = tk.Label(metrics_frame, text="LF/HF: --", font=metric_font, fg="white", bg="#1E1E1E")
        lbl_lfhf.grid(row=0, column=2, padx=8)
        lbl_beats = tk.Label(metrics_frame, text="Beats: --", font=metric_font, fg="white", bg="#1E1E1E")
        lbl_beats.grid(row=0, column=3, padx=8)
        lbl_window = tk.Label(metrics_frame, text="Window: -- s", font=metric_font, fg="white", bg="#1E1E1E")
        lbl_window.grid(row=0, column=4, padx=8)

        info_frame = tk.Frame(top_bar, bg="#1E1E1E")
        info_frame.pack(side="right", padx=6)
        btn_export = ttk.Button(info_frame, text="Export CSV",
                                command=lambda: self.export_hrv_csv_dialog(window_seconds_default=None))
        btn_export.pack(side="right", padx=(8, 0))
        btn_close = ttk.Button(info_frame, text="Close", command=win.destroy)
        btn_close.pack(side="right")

        fig = plt.Figure(figsize=(11, 7), dpi=100)
        fig.patch.set_facecolor("#1E1E1E")

        axes = fig.subplots(2, 2)
        axes = axes.flatten()
        fig.subplots_adjust(hspace=0.35, wspace=0.3)

        def _style_axis(ax):
            ax.set_facecolor("#1E1E1E")
            ax.tick_params(axis="x", colors="white")
            ax.tick_params(axis="y", colors="white")
            for spine in ax.spines.values():
                spine.set_color("#444444")
            ax.title.set_color("white")
            ax.xaxis.label.set_color("white")
            ax.yaxis.label.set_color("white")
            ax.grid(True, linestyle="--", alpha=0.25)

        for ax in axes:
            _style_axis(ax)

        canvas = FigureCanvasTkAgg(fig, master=win)
        canvas_widget = canvas.get_tk_widget()
        canvas_widget.pack(fill="both", expand=True, padx=10, pady=(4, 8))

        try:
            from matplotlib.backends.backend_tkagg import NavigationToolbar2Tk
            toolbar = NavigationToolbar2Tk(canvas, win)
            toolbar.update()
            toolbar.pack(fill="x", padx=10, pady=(0, 8))
        except Exception:
            pass

        note_label = tk.Label(win, text="", font=small_font, fg="#ffcc00", bg="#1E1E1E")
        note_label.pack(anchor="w", padx=12, pady=(0, 8))

        def update_plot():
            try:
                results = self.processor.compute_hrv(window_seconds=None)
            except Exception as e:
                results = {"error": str(e)}

            for ax in axes:
                ax.cla()
                _style_axis(ax)

            if not results or "error" in results:
                axes[0].text(0.5, 0.5, "Waiting for RR data...", ha="center", va="center", color="white",
                             transform=axes[0].transAxes)
                lbl_rmssd.config(text="RMSSD: -- ms")
                lbl_sdnn.config(text="SDNN: -- ms")
                lbl_lfhf.config(text="LF/HF: --")
                lbl_beats.config(text="Beats: --")
                lbl_window.config(text="Window: -- s")
                note_label.config(text="")
                canvas.draw_idle()
                win.after(max(500, int(getattr(self, "hrv_update_interval", 1.0) * 1000)), update_plot)
                return

            rr = np.asarray(results.get("rr_ms", []), dtype=float).reshape(-1)
            beats_used = int(results.get("beats_used", rr.size))
            actual_window_s = float(results.get("actual_window_s", (np.sum(rr) / 1000.0) if rr.size else 0.0))
            td = results.get("time_domain", {}) or {}
            poi = results.get("poincare", {}) or {}
            fd = results.get("freq_domain", {}) or {}

            rmssd = td.get("RMSSD_ms", "--")
            sdnn = td.get("SDNN_ms", "--")
            lf_hf = fd.get("LF_HF", "--") if fd else "--"

            lbl_rmssd.config(text=f"RMSSD: {rmssd} ms")
            lbl_sdnn.config(text=f"SDNN: {sdnn} ms")
            lbl_lfhf.config(text=f"LF/HF: {lf_hf}")
            lbl_beats.config(text=f"Beats: {beats_used}")
            lbl_window.config(text=f"Window: {round(actual_window_s, 2)} s")

            if actual_window_s > 0 and actual_window_s < 120:
                note_label.config(text="Spectrum may be less reliable for windows < 120 s")
            else:
                note_label.config(text="")

            if rr.size:
                t = np.cumsum(rr) / 1000.0
                axes[0].plot(t, rr, linewidth=1.5, color="cyan")
                axes[0].set_title("RR intervals (ms)")
                axes[0].set_xlabel("Time (s)")
                axes[0].set_ylabel("Interval (ms)")
                axes[0].grid(True, linestyle="--", alpha=0.3)
                axes[0].set_xlim(left=0, right=max(1.0, t[-1]))
                last_rr = rr[-1]
                axes[0].annotate(f"last: {int(last_rr)} ms", xy=(t[-1], last_rr), xytext=(-10, 8),
                                 textcoords="offset points", color="white", ha="right", va="bottom", fontsize=9)

            if rr.size > 1:
                axes[1].scatter(rr[:-1], rr[1:], s=18, alpha=0.8, color="lime")
                axes[1].set_title("Poincaré Plot")
                axes[1].set_xlabel("RR[n] (ms)")
                axes[1].set_ylabel("RR[n+1] (ms)")
                axes[1].plot([np.min(rr), np.max(rr)], [np.min(rr), np.max(rr)], color="#666666", linestyle="--",
                             linewidth=0.8)
                sd1 = poi.get("SD1_ms", None)
                sd2 = poi.get("SD2_ms", None)
                if sd1 is not None and sd2 is not None:
                    axes[1].text(0.98, 0.02, f"SD1={sd1} ms\nSD2={sd2} ms", transform=axes[1].transAxes, ha="right",
                                 va="bottom", color="white", fontsize=9)

            if rr.size:
                axes[2].hist(rr, bins=min(40, max(6, int(rr.size / 2))), alpha=0.85, edgecolor="#222222", linewidth=0.6)
                axes[2].set_title("Histogram of RR")
                axes[2].set_xlabel("Interval (ms)")
                axes[2].set_ylabel("Count")

            if fd and all(k in fd for k in ("VLF", "LF", "HF")):
                vals = [fd.get("VLF", 0.0), fd.get("LF", 0.0), fd.get("HF", 0.0)]
                labels = ["VLF", "LF", "HF"]
                colors = ["#ffaa00", "#00aaff", "#ff66cc"]
                bars = axes[3].bar(labels, vals, color=colors, edgecolor="#222222")
                axes[3].set_title("Frequency Bands Power")
                axes[3].set_ylabel("Power (a.u.)")
                axes[3].set_xlabel("")
                axes[3].grid(False)
                lfn = fd.get("LF_norm_pct", None)
                hfn = fd.get("HF_norm_pct", None)
                lf_hf_text = f"LF/HF: {fd.get('LF_HF', '--')}"
                if lfn is not None and hfn is not None:
                    lf_hf_text += f"  |  LF%: {lfn}  HF%: {hfn}"
                axes[3].text(0.98, 0.95, lf_hf_text, transform=axes[3].transAxes, ha="right", va="top", color="white",
                             fontsize=10)
                for bar in bars:
                    height = bar.get_height()
                    axes[3].annotate(f"{round(height, 3)}", xy=(bar.get_x() + bar.get_width() / 2, height),
                                     xytext=(0, 6), textcoords="offset points", ha="center", color="white", fontsize=9)
            else:
                axes[3].text(0.5, 0.5, "No frequency data", ha="center", va="center", color="white",
                             transform=axes[3].transAxes)

            canvas.draw_idle()
            win.after(max(200, int(getattr(self, "hrv_update_interval", 1.0) * 1000)), update_plot)

        update_plot()

    def export_hrv_csv_dialog(self, window_seconds_default=None):
        try:
            import numpy as np
            import datetime
            import os
            import tempfile
            import shutil
            from tkinter import filedialog, messagebox
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.drawing.image import Image as XLImage
            import matplotlib.pyplot as plt
            from scipy.signal import lombscargle

            # Professional gray theme colors
            THEME = {
                "header_bg": "2D2D2D",  # Dark gray header background
                "header_fg": "FFFFFF",  # White header text
                "section_bg": "3C3C3C",  # Slightly lighter gray for sections
                "section_fg": "E0E0E0",  # Light gray section text
                "data_bg": "252525",  # Darker gray for data cells
                "data_fg": "D0D0D0",  # Light gray data text
                "accent": "4A90E2",  # Blue accent color
                "warning": "FF6B6B",  # Red for warnings
                "success": "4CAF50"  # Green for success
            }

            # Professional fonts
            HEADER_FONT = Font(name="Segoe UI", size=16, bold=True, color=THEME["header_fg"])
            SECTION_FONT = Font(name="Segoe UI", size=14, bold=True, color=THEME["section_fg"])
            METRIC_FONT = Font(name="Segoe UI", size=12, bold=True, color=THEME["accent"])
            DATA_FONT = Font(name="Segoe UI", size=11, color=THEME["data_fg"])
            SMALL_FONT = Font(name="Segoe UI", size=10, color=THEME["data_fg"])

            # Cell styles
            HEADER_FILL = PatternFill("solid", fgColor=THEME["header_bg"])
            SECTION_FILL = PatternFill("solid", fgColor=THEME["section_bg"])
            DATA_FILL = PatternFill("solid", fgColor=THEME["data_bg"])
            BORDER = Border(
                left=Side(style="thin", color="555555"),
                right=Side(style="thin", color="555555"),
                top=Side(style="thin", color="555555"),
                bottom=Side(style="thin", color="555555")
            )
            CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
            LEFT_ALIGN = Alignment(horizontal="left", vertical="center", indent=1)

            # Chart height constants
            RR_CHART_ROWS = 20
            PSD_CHART_ROWS = 20
            SMALL_CHART_ROWS = 18

            # Parse window seconds
            window_seconds = window_seconds_default
            if hasattr(self, "entry_hrv_window") and self.entry_hrv_window:
                try:
                    val = self.entry_hrv_window.get().strip()
                    if val:
                        window_seconds = float(val)
                except (ValueError, TypeError):
                    pass

            # Get HRV data
            hrv = self.processor.compute_hrv(window_seconds=window_seconds)
            if "error" in hrv:
                messagebox.showinfo("Export HRV", f"Cannot export HRV: {hrv.get('error')}")
                return

            # Extract RR intervals
            rr = np.asarray(hrv.get("rr_ms", []), dtype=float)
            beats_used = len(rr)
            actual_window_s = float(np.sum(rr) / 1000.0) if beats_used > 0 else 0.0

            # Setup file paths
            base = os.path.basename(getattr(self.processor, "record_name", "ecg_record"))
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            default_name = f"hrv_report_{base}_{timestamp}.xlsx"

            try:
                rec_path = getattr(self.processor, "record_name", "")
                init_dir = os.path.dirname(rec_path) if rec_path and os.path.isdir(
                    os.path.dirname(rec_path)) else os.getcwd()
            except Exception:
                init_dir = os.getcwd()

            save_path = filedialog.asksaveasfilename(
                title="Save HRV Excel",
                initialdir=init_dir,
                initialfile=default_name,
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if not save_path:
                return

            # Create temporary directory for images (in same directory as save_path)
            img_dir = os.path.join(os.path.dirname(save_path), f"hrv_images_{timestamp}")
            os.makedirs(img_dir, exist_ok=True)

            # Create workbook with professional styling
            wb = Workbook()
            ws = wb.active
            ws.title = "HRV Report"

            # Set column widths as specified in the requirements
            ws.column_dimensions['A'].width = 42
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 2
            ws.column_dimensions['D'].width = 50
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 18

            # Set row heights for spacing
            ws.row_dimensions[1].height = 30  # Title row
            ws.row_dimensions[2].height = 20  # Subtitle row
            ws.row_dimensions[3].height = 10  # Spacer row

            # ========== TITLE SECTION ==========
            # Merge A1:F1 for title
            ws.merge_cells('A1:F1')
            title_cell = ws['A1']
            title_cell.value = "ECG Heart Rate Variability Analysis Report"
            title_cell.font = HEADER_FONT
            title_cell.fill = HEADER_FILL
            title_cell.alignment = CENTER_ALIGN
            title_cell.border = BORDER

            # Merge A2:F2 for subtitle
            ws.merge_cells('A2:F2')
            subtitle_cell = ws['A2']
            subtitle_cell.value = "Generated by ECG Signal Analysis Suite"
            subtitle_cell.font = SMALL_FONT
            subtitle_cell.fill = DATA_FILL
            subtitle_cell.alignment = CENTER_ALIGN
            subtitle_cell.border = BORDER

            # ========== METADATA SECTION ==========
            current_row = 4  # Start after title, subtitle, and spacer

            # Add section header
            ws.merge_cells(start_row=current_row, end_row=current_row, start_column=1, end_column=2)
            header_cell = ws.cell(row=current_row, column=1)
            header_cell.value = "Patient Information"
            header_cell.font = SECTION_FONT
            header_cell.fill = SECTION_FILL
            header_cell.alignment = LEFT_ALIGN
            header_cell.border = BORDER
            ws.row_dimensions[current_row].height = 28
            current_row += 1

            # Add metadata
            metadata = [
                ("Patient Name", getattr(self.processor, "patient_name", "Unknown")),
                ("Patient ID", getattr(self.processor, "patient_id", "Unknown")),
                ("Age", getattr(self.processor, "patient_age", "Unknown")),
                ("Gender", getattr(self.processor, "patient_gender", "Unknown")),
                ("Record File", base),
                ("Analysis Timestamp", timestamp),
                ("Analysis Window (s)", str(window_seconds) if window_seconds else "Entire Record"),
                ("Beats Used", beats_used),
                ("Actual Window Duration (s)", f"{actual_window_s:.3f}")
            ]

            for key, value in metadata:
                # Key column
                key_cell = ws.cell(row=current_row, column=1)
                key_cell.value = key
                key_cell.font = METRIC_FONT
                key_cell.fill = DATA_FILL
                key_cell.border = BORDER
                key_cell.alignment = LEFT_ALIGN

                # Value column
                value_cell = ws.cell(row=current_row, column=2)
                value_cell.value = value
                value_cell.font = DATA_FONT
                value_cell.fill = DATA_FILL
                value_cell.border = BORDER
                value_cell.alignment = LEFT_ALIGN

                ws.row_dimensions[current_row].height = 20
                current_row += 1

            # Add spacer row
            current_row += 1

            # ========== TIME-DOMAIN SECTION ==========
            # Add section header
            ws.merge_cells(start_row=current_row, end_row=current_row, start_column=1, end_column=2)
            header_cell = ws.cell(row=current_row, column=1)
            header_cell.value = "Time-Domain HRV Metrics"
            header_cell.font = SECTION_FONT
            header_cell.fill = SECTION_FILL
            header_cell.alignment = LEFT_ALIGN
            header_cell.border = BORDER
            ws.row_dimensions[current_row].height = 28
            current_row += 1

            # Add time-domain metrics
            td = hrv.get("time_domain", {})
            if not td:
                td = {
                    "meanNN_ms": "N/A",
                    "SDNN_ms": "N/A",
                    "RMSSD_ms": "N/A",
                    "pNN50_pct": "N/A"
                }

            time_domain_data = [
                ("Mean RR Interval", f"{td.get('meanNN_ms', 'N/A')} ms"),
                ("SDNN", f"{td.get('SDNN_ms', 'N/A')} ms"),
                ("RMSSD", f"{td.get('RMSSD_ms', 'N/A')} ms"),
                ("pNN50", f"{td.get('pNN50_pct', 'N/A')}%")
            ]

            for key, value in time_domain_data:
                # Key column
                key_cell = ws.cell(row=current_row, column=1)
                key_cell.value = key
                key_cell.font = METRIC_FONT
                key_cell.fill = DATA_FILL
                key_cell.border = BORDER
                key_cell.alignment = LEFT_ALIGN

                # Value column
                value_cell = ws.cell(row=current_row, column=2)
                value_cell.value = value
                value_cell.font = DATA_FONT
                value_cell.fill = DATA_FILL
                value_cell.border = BORDER
                value_cell.alignment = LEFT_ALIGN

                ws.row_dimensions[current_row].height = 20
                current_row += 1

            # Add RR plot
            anchor_rr = current_row
            try:
                if beats_used >= 2:
                    # Generate RR plot
                    plt.figure(figsize=(10, 6), facecolor="#252525")
                    plt.rcParams.update({'font.size': 12, 'axes.facecolor': '#252525',
                                         'axes.edgecolor': 'white', 'axes.labelcolor': 'white',
                                         'xtick.color': 'white', 'ytick.color': 'white'})

                    t = np.cumsum(rr) / 1000.0
                    plt.plot(t, rr, color="#4A90E2", linewidth=2.5)
                    plt.title("RR Intervals Over Time", fontsize=14, color="white", pad=15)
                    plt.xlabel("Time (seconds)", fontsize=12, color="white")
                    plt.ylabel("RR Interval (ms)", fontsize=12, color="white")
                    plt.grid(True, linestyle="--", alpha=0.3)
                    plt.tight_layout()

                    rr_plot_path = os.path.join(img_dir, "rr_intervals.png")
                    plt.savefig(rr_plot_path, dpi=150, bbox_inches="tight", facecolor="#252525")
                    plt.close()

                    # Embed in Excel
                    img = XLImage(rr_plot_path)
                    img.width, img.height = 750, 450
                    ws.add_image(img, f"D{anchor_rr}")
                else:
                    print(f"Skipping RR plot: Not enough beats ({beats_used} < 2)")
            except Exception as e:
                print(f"Skipping RR plot: Error generating plot - {str(e)}")

            # Advance current_row past the chart
            current_row = anchor_rr + RR_CHART_ROWS + 1

            # Add spacer row
            current_row += 1

            # ========== FREQUENCY-DOMAIN SECTION ==========
            # Add section header
            ws.merge_cells(start_row=current_row, end_row=current_row, start_column=1, end_column=2)
            header_cell = ws.cell(row=current_row, column=1)
            header_cell.value = "Frequency-Domain HRV Metrics"
            header_cell.font = SECTION_FONT
            header_cell.fill = SECTION_FILL
            header_cell.alignment = LEFT_ALIGN
            header_cell.border = BORDER
            ws.row_dimensions[current_row].height = 28
            current_row += 1

            # Add frequency-domain metrics
            fd = hrv.get("freq_domain", {})
            if not fd:
                fd = {
                    "VLF": "N/A",
                    "LF": "N/A",
                    "HF": "N/A",
                    "LF_norm_pct": "N/A",
                    "HF_norm_pct": "N/A",
                    "LF_HF": "N/A"
                }

            freq_domain_data = [
                ("Very Low Frequency (VLF)", f"{fd.get('VLF', 'N/A')} ms²"),
                ("Low Frequency (LF)", f"{fd.get('LF', 'N/A')} ms²"),
                ("High Frequency (HF)", f"{fd.get('HF', 'N/A')} ms²"),
                ("LF (Normalized)", f"{fd.get('LF_norm_pct', 'N/A')}%"),
                ("HF (Normalized)", f"{fd.get('HF_norm_pct', 'N/A')}%"),
                ("LF/HF Ratio", f"{fd.get('LF_HF', 'N/A')}")
            ]

            for key, value in freq_domain_data:
                # Key column
                key_cell = ws.cell(row=current_row, column=1)
                key_cell.value = key
                key_cell.font = METRIC_FONT
                key_cell.fill = DATA_FILL
                key_cell.border = BORDER
                key_cell.alignment = LEFT_ALIGN

                # Value column
                value_cell = ws.cell(row=current_row, column=2)
                value_cell.value = value
                value_cell.font = DATA_FONT
                value_cell.fill = DATA_FILL
                value_cell.border = BORDER
                value_cell.alignment = LEFT_ALIGN

                ws.row_dimensions[current_row].height = 20
                current_row += 1

            # Add PSD plot
            anchor_psd = current_row
            try:
                fd_full = hrv.get("freq_domain_full", {})
                freqs = None
                psd = None

                # Try to get PSD from freq_domain_full
                if isinstance(fd_full, dict) and "f" in fd_full and "psd" in fd_full:
                    freqs = np.asarray(fd_full["f"], dtype=float)
                    psd = np.asarray(fd_full["psd"], dtype=float)

                # Fallback: compute PSD using Lomb-Scargle
                if (freqs is None or psd is None) and beats_used >= 3:
                    t = np.cumsum(rr) / 1000.0
                    if len(t) > 1:
                        freqs = np.linspace(0.0033, 0.5, 2048)
                        ang = 2.0 * np.pi * freqs
                        y = rr - np.mean(rr)
                        psd = lombscargle(t, y, ang, precenter=True, normalize=True)

                # Generate PSD plot if we have data
                if freqs is not None and psd is not None and len(freqs) == len(psd) and beats_used >= 3:
                    plt.figure(figsize=(10, 6), facecolor="#252525")
                    plt.rcParams.update({'font.size': 12, 'axes.facecolor': '#252525',
                                         'axes.edgecolor': 'white', 'axes.labelcolor': 'white',
                                         'xtick.color': 'white', 'ytick.color': 'white'})

                    # Convert to dB/Hz for better visualization
                    psd_db = 10 * np.log10(psd + 1e-12)
                    plt.semilogx(freqs, psd_db, color="#9C27B0", linewidth=2.5)

                    # Add frequency bands
                    lf_band = fd.get("LF_band", [0.04, 0.15])
                    hf_band = fd.get("HF_band", [0.15, 0.4])
                    plt.axvspan(lf_band[0], lf_band[1], color="#2196F3", alpha=0.2)
                    plt.axvspan(hf_band[0], hf_band[1], color="#4CAF50", alpha=0.2)

                    plt.title("Heart Rate Variability Power Spectrum", fontsize=14, color="white", pad=15)
                    plt.xlabel("Frequency (Hz)", fontsize=12, color="white")
                    plt.ylabel("Power (dB/Hz)", fontsize=12, color="white")
                    plt.grid(True, which="both", linestyle="--", alpha=0.3)
                    plt.tight_layout()

                    psd_plot_path = os.path.join(img_dir, "psd.png")
                    plt.savefig(psd_plot_path, dpi=150, bbox_inches="tight", facecolor="#252525")
                    plt.close()

                    # Embed in Excel
                    img = XLImage(psd_plot_path)
                    img.width, img.height = 750, 450
                    ws.add_image(img, f"D{anchor_psd}")
                else:
                    print(f"Skipping PSD plot: Not enough beats ({beats_used} < 3) or no PSD data")
            except Exception as e:
                print(f"Skipping PSD plot: Error generating plot - {str(e)}")

            # Advance current_row past the chart
            current_row = anchor_psd + PSD_CHART_ROWS + 1

            # Add spacer row
            current_row += 1

            # ========== NONLINEAR SECTION ==========
            # Add section header
            ws.merge_cells(start_row=current_row, end_row=current_row, start_column=1, end_column=2)
            header_cell = ws.cell(row=current_row, column=1)
            header_cell.value = "Nonlinear HRV Metrics"
            header_cell.font = SECTION_FONT
            header_cell.fill = SECTION_FILL
            header_cell.alignment = LEFT_ALIGN
            header_cell.border = BORDER
            ws.row_dimensions[current_row].height = 28
            current_row += 1

            # Add nonlinear metrics
            nonlinear_data = [
                ("Sample Entropy", hrv.get("sample_entropy", "N/A")),
                ("DFA Alpha", hrv.get("dfa_alpha", "N/A")),
                ("DFA Fluctuation", hrv.get("dfa_fluct", "N/A"))
            ]

            for key, value in nonlinear_data:
                # Key column
                key_cell = ws.cell(row=current_row, column=1)
                key_cell.value = key
                key_cell.font = METRIC_FONT
                key_cell.fill = DATA_FILL
                key_cell.border = BORDER
                key_cell.alignment = LEFT_ALIGN

                # Value column
                value_cell = ws.cell(row=current_row, column=2)
                value_cell.value = value
                value_cell.font = DATA_FONT
                value_cell.fill = DATA_FILL
                value_cell.border = BORDER
                value_cell.alignment = LEFT_ALIGN

                ws.row_dimensions[current_row].height = 20
                current_row += 1

            # Add Poincaré plot
            anchor_poincare = current_row
            try:
                poi = hrv.get("poincare", {})
                if beats_used > 1 and poi:
                    plt.figure(figsize=(8, 8), facecolor="#252525")
                    plt.rcParams.update({'font.size': 12, 'axes.facecolor': '#252525',
                                         'axes.edgecolor': 'white', 'axes.labelcolor': 'white',
                                         'xtick.color': 'white', 'ytick.color': 'white'})

                    plt.scatter(rr[:-1], rr[1:], s=35, alpha=0.8, color="#FF9800", edgecolor="none")
                    sd1 = poi.get("SD1_ms", 0)
                    sd2 = poi.get("SD2_ms", 0)

                    if sd1 > 0 and sd2 > 0:
                        # Draw SD1 and SD2 lines
                        mean_rr = np.mean(rr)
                        plt.plot([mean_rr - sd2 / 2, mean_rr + sd2 / 2],
                                 [mean_rr - sd2 / 2, mean_rr + sd2 / 2], 'b--', alpha=0.5)
                        plt.plot([mean_rr, mean_rr + sd1],
                                 [mean_rr, mean_rr - sd1], 'r--', alpha=0.5)

                    plt.title("Poincaré Plot", fontsize=14, color="white", pad=15)
                    plt.xlabel("RR[n] (ms)", fontsize=12, color="white")
                    plt.ylabel("RR[n+1] (ms)", fontsize=12, color="white")
                    plt.grid(True, linestyle="--", alpha=0.3)
                    plt.axis('equal')
                    plt.tight_layout()

                    poincare_plot_path = os.path.join(img_dir, "poincare.png")
                    plt.savefig(poincare_plot_path, dpi=150, bbox_inches="tight", facecolor="#252525")
                    plt.close()

                    # Embed in Excel
                    img = XLImage(poincare_plot_path)
                    img.width, img.height = 600, 600
                    ws.add_image(img, f"D{anchor_poincare}")
                else:
                    print(f"Skipping Poincaré plot: Not enough beats ({beats_used} <= 1)")
            except Exception as e:
                print(f"Skipping Poincaré plot: Error generating plot - {str(e)}")

            # Advance current_row past the chart
            current_row = anchor_poincare + SMALL_CHART_ROWS + 1

            # Add DFA plot
            anchor_dfa = current_row
            try:
                if beats_used >= 20 and "dfa_fluct" in hrv and hrv["dfa_fluct"] is not None:
                    x = rr.astype(float)
                    N = len(x)
                    y = np.cumsum(x - np.mean(x))
                    scales = np.unique(np.floor(np.logspace(np.log10(4), np.log10(max(8, N // 4)), num=20))).astype(int)
                    F = []
                    for s in scales:
                        segments = N // s
                        if segments < 2:
                            F.append(np.nan)
                            continue
                        rms = []
                        for i in range(segments):
                            seg = y[i * s:(i + 1) * s]
                            if len(seg) < 2:
                                continue
                            tvec = np.arange(len(seg))
                            try:
                                coeffs = np.polyfit(tvec, seg, 1)
                                trend = np.polyval(coeffs, tvec)
                                rms_val = np.sqrt(np.mean((seg - trend) ** 2))
                                rms.append(rms_val)
                            except Exception:
                                continue
                        F.append(np.mean(rms) if rms else np.nan)

                    # Filter out NaN values
                    valid_idx = [i for i, f in enumerate(F) if not np.isnan(f)]
                    if len(valid_idx) >= 3:  # Need at least 3 points for meaningful plot
                        scales = [scales[i] for i in valid_idx]
                        F = [F[i] for i in valid_idx]

                        plt.figure(figsize=(10, 6), facecolor="#252525")
                        plt.rcParams.update({'font.size': 12, 'axes.facecolor': '#252525',
                                             'axes.edgecolor': 'white', 'axes.labelcolor': 'white',
                                             'xtick.color': 'white', 'ytick.color': 'white'})

                        plt.loglog(scales, F, 'o-', color="#E91E63", markersize=6, linewidth=2)

                        # Calculate DFA alpha if possible
                        alpha = np.polyfit(np.log10(scales), np.log10(F), 1)[0]
                        plt.text(0.05, 0.95, f"DFA Alpha: {alpha:.3f}",
                                 transform=plt.gca().transAxes,
                                 fontsize=12, color="white",
                                 bbox=dict(facecolor='black', alpha=0.5))

                        plt.title("Detrended Fluctuation Analysis", fontsize=14, color="white", pad=15)
                        plt.xlabel("Scale", fontsize=12, color="white")
                        plt.ylabel("F(s)", fontsize=12, color="white")
                        plt.grid(True, which="both", linestyle="--", alpha=0.3)
                        plt.tight_layout()

                        dfa_plot_path = os.path.join(img_dir, "dfa.png")
                        plt.savefig(dfa_plot_path, dpi=150, bbox_inches="tight", facecolor="#252525")
                        plt.close()

                        # Embed in Excel
                        img = XLImage(dfa_plot_path)
                        img.width, img.height = 750, 450
                        ws.add_image(img, f"D{anchor_dfa}")
                    else:
                        print(f"Skipping DFA plot: Not enough valid points ({len(valid_idx)} < 3)")
                else:
                    print(f"Skipping DFA plot: Not enough beats ({beats_used} < 20)")
            except Exception as e:
                print(f"Skipping DFA plot: Error generating plot - {str(e)}")

            # Advance current_row past the chart
            current_row = anchor_dfa + SMALL_CHART_ROWS + 1

            # Add spacer row
            current_row += 1

            # ========== MSE SECTION ==========
            # Add section header
            ws.merge_cells(start_row=current_row, end_row=current_row, start_column=1, end_column=2)
            header_cell = ws.cell(row=current_row, column=1)
            header_cell.value = "Multi-scale Entropy"
            header_cell.font = SECTION_FONT
            header_cell.fill = SECTION_FILL
            header_cell.alignment = LEFT_ALIGN
            header_cell.border = BORDER
            ws.row_dimensions[current_row].height = 28
            current_row += 1

            # Add MSE metrics
            mse = hrv.get("multi_scale_entropy", [])
            valid_mse = [x for x in mse if not np.isnan(x)] if mse else []

            mse_data = [
                ("MSE Mean", f"{np.mean(valid_mse):.3f}" if valid_mse else "N/A"),
                ("MSE Scales Count", len(valid_mse)),
                ("Sample Entropy", hrv.get("sample_entropy", "N/A"))
            ]

            for key, value in mse_data:
                # Key column
                key_cell = ws.cell(row=current_row, column=1)
                key_cell.value = key
                key_cell.font = METRIC_FONT
                key_cell.fill = DATA_FILL
                key_cell.border = BORDER
                key_cell.alignment = LEFT_ALIGN

                # Value column
                value_cell = ws.cell(row=current_row, column=2)
                value_cell.value = value
                value_cell.font = DATA_FONT
                value_cell.fill = DATA_FILL
                value_cell.border = BORDER
                value_cell.alignment = LEFT_ALIGN

                ws.row_dimensions[current_row].height = 20
                current_row += 1

            # Add MSE plot
            anchor_mse = current_row
            try:
                if valid_mse and len(valid_mse) >= 2:
                    plt.figure(figsize=(10, 6), facecolor="#252525")
                    plt.rcParams.update({'font.size': 12, 'axes.facecolor': '#252525',
                                         'axes.edgecolor': 'white', 'axes.labelcolor': 'white',
                                         'xtick.color': 'white', 'ytick.color': 'white'})

                    plt.plot(range(1, len(valid_mse) + 1), valid_mse, marker="o", color="#4CAF50",
                             markersize=6, linewidth=2.5)

                    # Add average line
                    avg_mse = np.mean(valid_mse)
                    plt.axhline(y=avg_mse, color="#F44336", linestyle="--", alpha=0.7)
                    plt.text(len(valid_mse) * 0.95, avg_mse * 1.05, f"Mean: {avg_mse:.3f}",
                             color="#F44336", fontsize=10)

                    plt.title("Multi-scale Entropy Analysis", fontsize=14, color="white", pad=15)
                    plt.xlabel("Scale", fontsize=12, color="white")
                    plt.ylabel("Sample Entropy", fontsize=12, color="white")
                    plt.grid(True, linestyle="--", alpha=0.3)
                    plt.tight_layout()

                    mse_plot_path = os.path.join(img_dir, "mse.png")
                    plt.savefig(mse_plot_path, dpi=150, bbox_inches="tight", facecolor="#252525")
                    plt.close()

                    # Embed in Excel
                    img = XLImage(mse_plot_path)
                    img.width, img.height = 750, 450
                    ws.add_image(img, f"D{anchor_mse}")
                else:
                    print(f"Skipping MSE plot: Not enough valid MSE values ({len(valid_mse)} < 2)")
            except Exception as e:
                print(f"Skipping MSE plot: Error generating plot - {str(e)}")

            # Advance current_row past the chart
            current_row = anchor_mse + SMALL_CHART_ROWS + 1

            # Add spacer row
            current_row += 1

            # ========== RR DATA SECTION ==========
            # Add section header
            ws.merge_cells(start_row=current_row, end_row=current_row, start_column=1, end_column=2)
            header_cell = ws.cell(row=current_row, column=1)
            header_cell.value = "RR Intervals Data"
            header_cell.font = SECTION_FONT
            header_cell.fill = SECTION_FILL
            header_cell.alignment = LEFT_ALIGN
            header_cell.border = BORDER
            ws.row_dimensions[current_row].height = 28
            current_row += 1

            # Add RR data table header
            ws.cell(row=current_row, column=1).value = "Beat Index"
            ws.cell(row=current_row, column=2).value = "RR Interval (ms)"

            # Style header
            for col in range(1, 3):
                cell = ws.cell(row=current_row, column=col)
                cell.font = METRIC_FONT
                cell.fill = SECTION_FILL
                cell.border = BORDER
                cell.alignment = CENTER_ALIGN

            # Advance to data rows
            current_row += 1

            # Add RR data
            for i, val in enumerate(rr):
                ws.cell(row=current_row, column=1).value = i + 1
                ws.cell(row=current_row, column=2).value = float(val)

                # Style data cells
                for col in range(1, 3):
                    cell = ws.cell(row=current_row, column=col)
                    cell.font = DATA_FONT
                    cell.fill = DATA_FILL
                    cell.border = BORDER
                    cell.alignment = CENTER_ALIGN

                current_row += 1

            # Save the workbook
            wb.save(save_path)

            # Clean up temporary images
            try:
                shutil.rmtree(img_dir)
            except Exception as e:
                print(f"Warning: Could not clean up temporary image directory: {str(e)}")

            messagebox.showinfo("Export HRV", f"HRV report generated successfully!\n\n"
                                              f"File saved to:\n{save_path}")

        except Exception as e:
            # Clean up temporary images in case of error
            try:
                if 'img_dir' in locals() and os.path.exists(img_dir):
                    shutil.rmtree(img_dir)
            except:
                pass

            messagebox.showerror("Export HRV Error", f"Failed to export HRV:\n{str(e)}")

    def run(self):
        self.root.mainloop()


def login_page():
    global root_login
    root_login = tk.Tk()
    root_login.title("ECG Signal Login")
    root_login.geometry("420x280")
    root_login.configure(bg="#1E1E1E")

    style = ttk.Style()
    style.configure("TLabel", font=("Segoe UI", 11), background="#1E1E1E", foreground="white")
    style.configure("TEntry", font=("Segoe UI", 11))
    style.configure("TButton", font=("Segoe UI", 12, "bold"), padding=6)
    style.map("TButton",
              foreground=[("active", "black")],
              background=[("active", "#1BA1E2")])

    title = tk.Label(root_login, text="🔐 ECG Signal Loader", font=("Segoe UI", 16, "bold"),
                     bg="#1E1E1E", fg="cyan")
    title.pack(pady=15)

    lbl_path = ttk.Label(root_login, text="Enter Record Path or Browse:")
    lbl_path.pack(pady=5)
    entry_path = ttk.Entry(root_login, width=35)
    entry_path.pack(pady=5)

    def browse_file():
        file_path = filedialog.askopenfilename(
            title="Select ECG File (.dat or .hea)",
            filetypes=[("WFDB Files", "*.dat *.hea"), ("All Files", "*.*")]
        )
        if file_path:
            base_name = os.path.splitext(file_path)[0]
            if not os.path.exists(base_name + ".dat"):
                messagebox.showerror("Error", "Missing .dat file!")
                return
            entry_path.delete(0, tk.END)
            entry_path.insert(0, base_name)

    btn_browse = ttk.Button(root_login, text="Browse", command=browse_file)
    btn_browse.pack(pady=5)
    def start_clicked():
        path = entry_path.get().strip()
        if not path:
            messagebox.showerror("Error", "Please enter a valid path or record number.")
            return

        base_name = os.path.splitext(path)[0]
        if path.isnumeric():
            base_name = path

        try:
            wfdb.rdrecord(base_name)
        except Exception as e:
            dat_path = base_name + ".dat"
            if not os.path.exists(dat_path):
                messagebox.showerror("Error", f"Record not found: {base_name}\nCheck path or use built-in like '100'.")
                return

        root_login.destroy()
        start_main_app(base_name)

    btn_start = ttk.Button(root_login, text="Start Processing", command=start_clicked)
    btn_start.pack(pady=20, ipadx=10, ipady=5)

    root_login.mainloop()


if __name__ == "__main__":
    login_page()